from unidecode import unidecode
import random
from openpyxl import Workbook
from faker import Faker
from faker.providers.phone_number import Provider as PhoneNumberProvider

import pandas as pd

df = pd.read_csv('grupos.csv')


book = Workbook()
ws = book.active

fake = Faker('pt_BR')

class CustomBrazilProvider(PhoneNumberProvider):
    formats = (
        '+55 (011) #### ####',
        '+55 (021) #### ####',
        '+55 (031) #### ####',
        '+55 (041) #### ####',
        '+55 (051) #### ####',
        '+55 (061) #### ####',
        '+55 (071) #### ####',
        '+55 (081) #### ####',
        '+55 (084) #### ####',
        '+55 11 #### ####',
        '+55 21 #### ####',
        '+55 31 #### ####',
        '+55 41 #### ####',
        '+55 51 #### ####',
        '+55 61 #### ####',
        '+55 71 #### ####',
        '+55 81 #### ####',
        '+55 84 #### ####',
        '+55 (011) ####-####',
        '+55 (021) ####-####',
        '+55 (031) ####-####',
        '+55 (041) ####-####',
        '+55 (051) ####-####',
        '+55 (061) ####-####',
        '+55 (071) ####-####',
        '+55 (081) ####-####',
        '+55 (084) ####-####',
        '+55 11 ####-####',
        '+55 21 ####-####',
        '+55 31 ####-####',
        '+55 41 ####-####',
        '+55 51 #### ####',
        '+55 61 ####-####',
        '+55 71 ####-####',
        '+55 81 ####-####',
        '+55 84 ####-####',
        '(011) #### ####',
        '(021) #### ####',
        '(031) #### ####',
        '(041) #### ####',
        '(051) #### ####',
        '(061) #### ####',
        '(071) #### ####',
        '(081) #### ####',
        '(084) #### ####',
        '11 #### ####',
        '21 #### ####',
        '31 #### ####',
        '41 #### ####',
        '51 #### ####',
        '61 #### ####',
        '71 #### ####',
        '81 #### ####',
        '84 #### ####',
        '(011) ####-####',
        '(021) ####-####',
        '(031) ####-####',
        '(041) ####-####',
        '(051) ####-####',
        '(061) ####-####',
        '(071) ####-####',
        '(081) ####-####',
        '(084) ####-####',
        '11 ####-####',
        '21 ####-####',
        '31 ####-####',
        '41 ####-####',
        '51 #### ####',
        '61 ####-####',
        '71 ####-####',
        '81 ####-####',
        '84 ####-####',
    )

    msisdn_formats = (
        '55119########',
        '55219########',
        '55319########',
        '55419########',
        '55519########',
        '55619########',
        '55719########',
        '55819########',
        '55849########',
    )

    cellphone_formats = (
        '+55 1# 9#### ####',
        '+55 2# 9#### ####',
        '+55 3# 9#### ####',
        '+55 4# 9#### ####',
        '+55 5# 9#### ####',
        '+55 6# 9#### ####',
        '+55 7# 9#### ####',
        '+55 8# 9#### ####',
        '+55 1# 9 #### ####',
        '+55 2# 9 #### ####',
        '+55 3# 9 #### ####',
        '+55 4# 9 #### ####',
        '+55 5# 9 #### ####',
        '+55 6# 9 #### ####',
        '+55 7# 9 #### ####',
        '+55 8# 9 #### ####',
        '+55 (01#) 9#### ####',
        '+55 (02#) 9#### ####',
        '+55 (03#) 9#### ####',
        '+55 (04#) 9#### ####',
        '+55 (05#) 9#### ####',
        '+55 (06#) 9#### ####',
        '+55 (07#) 9#### ####',
        '+55 (08#) 9#### ####',
        '+55 (1#) 9#### ####',
        '+55 (2#) 9#### ####',
        '+55 (3#) 9#### ####',
        '+55 (4#) 9#### ####',
        '+55 (5#) 9#### ####',
        '+55 (6#) 9#### ####',
        '+55 (7#) 9#### ####',
        '+55 (8#) 9#### ####',
        '+55 (1#) 9 #### ####',
        '+55 (2#) 9 #### ####',
        '+55 (3#) 9 #### ####',
        '+55 (4#) 9 #### ####',
        '+55 (5#) 9 #### ####',
        '+55 (6#) 9 #### ####',
        '+55 (7#) 9 #### ####',
        '+55 (8#) 9 #### ####',
        '+55 1# 9####-####',
        '+55 2# 9####-####',
        '+55 3# 9####-####',
        '+55 4# 9####-####',
        '+55 5# 9####-####',
        '+55 6# 9####-####',
        '+55 7# 9####-####',
        '+55 8# 9####-####',
        '+55 1# 9 ####-####',
        '+55 2# 9 ####-####',
        '+55 3# 9 ####-####',
        '+55 4# 9 ####-####',
        '+55 5# 9 ####-####',
        '+55 6# 9 ####-####',
        '+55 7# 9 ####-####',
        '+55 8# 9 ####-####',
        '+55 (01#) 9####-####',
        '+55 (02#) 9####-####',
        '+55 (03#) 9####-####',
        '+55 (04#) 9####-####',
        '+55 (05#) 9####-####',
        '+55 (06#) 9####-####',
        '+55 (07#) 9####-####',
        '+55 (08#) 9####-####',
        '+55 (1#) 9####-####',
        '+55 (2#) 9####-####',
        '+55 (3#) 9####-####',
        '+55 (4#) 9####-####',
        '+55 (5#) 9####-####',
        '+55 (6#) 9####-####',
        '+55 (7#) 9####-####',
        '+55 (8#) 9####-####',
        '+55 (1#) 9 ####-####',
        '+55 (2#) 9 ####-####',
        '+55 (3#) 9 ####-####',
        '+55 (4#) 9 ####-####',
        '+55 (5#) 9 ####-####',
        '+55 (6#) 9 ####-####',
        '+55 (7#) 9 ####-####',
        '+55 (8#) 9 ####-####',
    )

    services_phones_formats = (
        '100',
        '128',
        '151',
        '152',
        '153',
        '156',
        '180',
        '181',
        '185',
        '188',
        '190',
        '191',
        '192',
        '193',
        '194',
        '197',
        '198',
        '199',
    )

    def brazil_cellphone_number(self):
        pattern = self.random_element(self.cellphone_formats)
        return self.numerify(self.generator.parse(pattern))

    def brazil_service_phone_number(self):
        pattern = self.random_element(self.services_phones_formats)
        return self.numerify(self.generator.parse(pattern))

fake.add_provider(CustomBrazilProvider)

cols = ["NOME", "SOBRENOME", "EMAIL", "TELEFONE", "CPF", "DATA_NASCIMENTO", "ENDERECO_LOGRADOURO",
        "ENDERECO_NUMERO", "ENDERECO_COMPLEMENTO", "ENDERECO_BAIRRO", "ENDERECO_CEP", "ENDERECO_CIDADE", "ENDERECO_ESTADO",
        "CARGO", "TIME", "USAR_ENDERECO_EMPRESA", "EMITIR_CARTAO"]
        
required_fields = ["(Obrigatório)", "(Obrigatório)", "(Obrigatório)", "(Opcional)","(Obrigatório, com pontos e traços)", "(Obrigatório, formato 'DD/MM/AAAA')", "(Obrigatório)",
                    "(Obrigatório)", "(Opcional)", "(Opcional)", "(Opcional)", "(Opcional)","(Opcional)",
                    "(Opcional)", "(Opcional)", "(Obrigatório, formato 'S' ou 'N')", "(Obrigatório, formato 'S' ou 'N')"]

unidades_federativas = {
    "TO": "Tocantins",
	"SP": "São Paulo",
	"SE": "Sergipe",
	"SC": "Santa Catarina",
	"RR": "Roraima",
	"RO": "Rondônia",
	"RS": "Rio Grande do Sul",
	"RN": "Rio Grande do Norte",
	"RJ": "Rio de Janeiro",
	"PI": "Piauí",
	"PE": "Pernambuco",
	"PA": "Pará",
	"PB": "Paraíba",
	"PR": "Paraná",
	"MG": "Minas Gerais",
	"MS": "Mato Grosso do Sul",
	"MT": "Mato Grosso",
	"MA": "Maranhão",
    "GO": "Goiás",
    "ES": "Espírito Santo",
    "DF": "Distrito Federal",
    "CE": "Ceará",
    "BA": "Bahia",
    "AM": "Amazonas",
    "AP": "Amapá",
    "AL": "Alagoas",
    "AC": "Acre"
}

for index, col in enumerate(cols):
    ws.cell(column=index + 1, row=1, value="{0}".format(col))
count = 1
persons = []

while count <= 800:
    person = []
    nome = ''
    for i in df.sample(n=count)['name'].str.capitalize():
        nome = i
    sobrenome = fake.last_name()
    person.append(nome)
    person.append(sobrenome)
    person.append(unidecode('.'.join(nome.lower().strip().split(' '))) + "." + unidecode('.'.join(sobrenome.lower().strip().split(' '))) + "@" + fake.domain_name())
    person.append(fake.brazil_cellphone_number())
    person.append(fake.cpf())
    person.append(fake.date_of_birth().strftime('%d/%m/%Y'))
    person.append(fake.street_address().split(',')[0])
    person.append(fake.building_number())
    person.append('') # Complemento
    person.append('Centro') # Bairro
    person.append(fake.postcode())
    person.append(fake.city())
    person.append(random.choice(list(unidades_federativas.keys())))
    person.append('') # Cargo
    person.append('') # Time
    person.append('N')
    person.append('S') #EMITIR CARTAO

    person_exist = [pers for pers in persons if pers[0] == person[0]]

    if len(person_exist) > 3:
        continue

    for index, col in enumerate(person):
        ws.cell(column=index + 1, row=count + 1, value="{0}".format(col))
    count += 1
    persons.append(person)

book.save(filename="TesteMassaColaboradores.xlsx")
